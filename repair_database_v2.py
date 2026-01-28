#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script para reparar inconsistencias en database.xlsx
Usa openpyxl para modificar sin perder datos
"""

import pandas as pd
import openpyxl
from datetime import datetime
import os

DB_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.xlsx')

print("=" * 70)
print("REPARACIÓN DE INCONSISTENCIAS EN DATABASE.XLSX")
print("=" * 70)

# Backup del archivo original
backup_file = DB_FILE.replace('.xlsx', '_backup.xlsx')
if not os.path.exists(backup_file):
    import shutil
    shutil.copy(DB_FILE, backup_file)
    print(f"\n✓ Backup creado: {backup_file}")
else:
    print(f"\n✓ Backup ya existe")

# 1. Reparar Inventario - Fecha_Actualizacion vacía
print("\n" + "=" * 70)
print("1. REPARANDO: Inventario - Fecha_Actualizacion vacía")
print("=" * 70)

df_inventario = pd.read_excel(DB_FILE, sheet_name='Inventario')
print(f"Valores nulos en Fecha_Actualizacion: {df_inventario['Fecha_Actualizacion'].isna().sum()}")

# Llenar con fecha actual solo si está vacío
mask = df_inventario['Fecha_Actualizacion'].isna()
df_inventario.loc[mask, 'Fecha_Actualizacion'] = datetime.now()
print(f"Valores nulos después de reparación: {df_inventario['Fecha_Actualizacion'].isna().sum()}")

# 2. Reparar Productos - NaN en Marca y Descripcion
print("\n" + "=" * 70)
print("2. REPARANDO: Productos - Marca y Descripcion vacías")
print("=" * 70)

df_productos = pd.read_excel(DB_FILE, sheet_name='Productos')
marca_nulos = df_productos['Marca'].isna().sum()
desc_nulos = df_productos['Descripcion'].isna().sum()
print(f"Marca - Valores nulos antes: {marca_nulos}")
print(f"Descripcion - Valores nulos antes: {desc_nulos}")

df_productos['Marca'] = df_productos['Marca'].fillna('SIN ESPECIFICAR')
df_productos['Descripcion'] = df_productos['Descripcion'].fillna('SIN DESCRIPCIÓN')

print(f"Marca - Valores nulos después: {df_productos['Marca'].isna().sum()}")
print(f"Descripcion - Valores nulos después: {df_productos['Descripcion'].isna().sum()}")

# 3. Reparar GastosMensuales - Convertir Fecha_Registro a datetime
print("\n" + "=" * 70)
print("3. REPARANDO: GastosMensuales - Fecha_Registro tipo de dato")
print("=" * 70)

df_gastos_mensuales = pd.read_excel(DB_FILE, sheet_name='GastosMensuales')
print(f"Tipo de dato Fecha_Registro: {df_gastos_mensuales['Fecha_Registro'].dtype}")

df_gastos_mensuales['Fecha_Registro'] = pd.to_datetime(df_gastos_mensuales['Fecha_Registro'])
print(f"Tipo de dato después: {df_gastos_mensuales['Fecha_Registro'].dtype}")

# 4. Verificar fechas sospechosas
print("\n" + "=" * 70)
print("4. VERIFICANDO: Fechas sospechosas")
print("=" * 70)

today = datetime.now()
problemas_encontrados = False

for sheet_name in ['Servicios', 'Productos', 'Gastos', 'Citas', 'Inventario', 'GastosMensuales']:
    if sheet_name == 'Inventario':
        df = df_inventario
    elif sheet_name == 'Productos':
        df = df_productos
    elif sheet_name == 'GastosMensuales':
        df = df_gastos_mensuales
    else:
        df = pd.read_excel(DB_FILE, sheet_name=sheet_name)
    
    # Encontrar columna de fecha
    date_col = None
    for col in ['Fecha', 'Fecha_Actualizacion', 'Fecha_Registro']:
        if col in df.columns:
            date_col = col
            break
    
    if date_col:
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
        future = (df[date_col] > today).sum()
        ancient = (df[date_col] < pd.Timestamp('2020-01-01')).sum()
        
        print(f"\n{sheet_name}: Total={len(df)}, Futuras={future}, Antiguas={ancient}", end="")
        if future > 0 or ancient > 0:
            problemas_encontrados = True
            print(" ⚠️ REVISAR")
        else:
            print(" ✓")

# Guardar cambios
print("\n" + "=" * 70)
print("GUARDANDO CAMBIOS")
print("=" * 70)

# Usar openpyxl para reemplazar cada sheet
wb = openpyxl.load_workbook(DB_FILE)

# Convertir DataFrames a Excel y actualizar sheets
dataframes_to_update = {
    'Inventario': df_inventario,
    'Productos': df_productos,
    'GastosMensuales': df_gastos_mensuales
}

for sheet_name, df in dataframes_to_update.items():
    # Eliminar sheet existente
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    
    # Agregar nuevo sheet
    ws = wb.create_sheet(sheet_name)
    
    # Escribir headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Escribir datos
    for row_idx, (idx, row) in enumerate(df.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

wb.save(DB_FILE)
print("\n✓ Cambios guardados en database.xlsx")

print("\n" + "=" * 70)
print("RESUMEN DE REPARACIONES")
print("=" * 70)
print(f"""
1. ✓ Inventario.Fecha_Actualizacion - {df_inventario['Fecha_Actualizacion'].isna().sum()} nulos
2. ✓ Productos.Marca - {df_productos['Marca'].isna().sum()} nulos
3. ✓ Productos.Descripcion - {df_productos['Descripcion'].isna().sum()} nulos
4. ✓ GastosMensuales.Fecha_Registro - Tipo datetime64
5. ✓ Verificación de fechas: {'⚠️ Encontrados problemas' if problemas_encontrados else '✓ OK'}

La aplicación debería cargar correctamente ahora.
Backup disponible en: {backup_file}
""")

print("=" * 70)
