from flask import Flask, render_template, request, jsonify, session, redirect, url_for, make_response
from werkzeug.security import generate_password_hash, check_password_hash
import functools
import pandas as pd
import os
from datetime import datetime
import openpyxl
import json
from io import BytesIO
from xhtml2pdf import pisa

app = Flask(__name__)
app.secret_key = 'magical_hair_secret_key_change_this_in_production'  # Required for session

# Use relative path for database.xlsx
DB_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.xlsx')
STYLISTS_FILE = 'stylists.json'
SERVICES_FILE = 'services.json'
SEDES_FILE = 'sedes.json'
USERS_FILE = 'users.json'

@app.errorhandler(500)
def internal_error(error):
    if request.path.startswith('/api/'):
        return jsonify({'status': 'error', 'message': 'Error interno del servidor. Verifique los logs.'}), 500
    return "Error interno del servidor", 500

@app.errorhandler(404)
def not_found_error(error):
    if request.path.startswith('/api/'):
        return jsonify({'status': 'error', 'message': 'Recurso no encontrado'}), 404
    return "Página no encontrada", 404

def login_required(view):
    @functools.wraps(view)
    def wrapped_view(**kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'status': 'error', 'message': 'Sesión expirada. Por favor recargue la página.'}), 401
            return redirect(url_for('login'))
        return view(**kwargs)
    return wrapped_view

def get_users():
    try:
        if not os.path.exists(USERS_FILE):
            # Create default admin user
            default_users = {
                'admin': generate_password_hash('admin')
            }
            save_users(default_users)
            return default_users
        with open(USERS_FILE, 'r') as f:
            return json.load(f)
    except:
        return {}

def save_users(users):
    with open(USERS_FILE, 'w') as f:
        json.dump(users, f, indent=2)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.json
        username = data.get('username')
        password = data.get('password')
        
        users = get_users()
        
        if username in users and check_password_hash(users[username], password):
            session.clear()
            session['user_id'] = username
            return jsonify({'status': 'success', 'message': 'Login exitoso'})
        
        return jsonify({'status': 'error', 'message': 'Usuario o contraseña incorrectos'})
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/api/users', methods=['GET'])
@login_required
def get_users_api():
    users = get_users()
    # Return list of usernames only
    return jsonify({'status': 'success', 'data': list(users.keys())})

@app.route('/api/users', methods=['POST'])
@login_required
def add_user():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'status': 'error', 'message': 'Usuario y contraseña requeridos'})
        
    users = get_users()
    if username in users:
        return jsonify({'status': 'error', 'message': 'El usuario ya existe'})
        
    users[username] = generate_password_hash(password)
    save_users(users)
    return jsonify({'status': 'success', 'message': 'Usuario creado'})

@app.route('/api/users', methods=['DELETE'])
@login_required
def delete_user():
    data = request.json
    username = data.get('username')
    
    if username == 'admin':
        return jsonify({'status': 'error', 'message': 'No se puede eliminar el usuario admin'})
        
    if username == session.get('user_id'):
        return jsonify({'status': 'error', 'message': 'No puedes eliminar tu propio usuario'})
        
    users = get_users()
    if username in users:
        del users[username]
        save_users(users)
        return jsonify({'status': 'success', 'message': 'Usuario eliminado'})
        
    return jsonify({'status': 'error', 'message': 'Usuario no encontrado'})



def check_and_migrate_db():
    """Check if the database has all required columns and add them if missing."""
    if os.path.exists(DB_FILE):
        try:
            # Check Services sheet
            df_services = pd.read_excel(DB_FILE, sheet_name='Servicios')
            needs_update = False
            
            if 'Metodo_Pago' not in df_services.columns:
                print("Migrating Services sheet: Adding Metodo_Pago column...")
                df_services['Metodo_Pago'] = 'Efectivo'
                needs_update = True
                
            if 'Sede' not in df_services.columns:
                print("Migrating Services sheet: Adding Sede column...")
                df_services.insert(0, 'Sede', 'Principal')
                needs_update = True
                
            if needs_update:
                with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df_services.to_excel(writer, sheet_name='Servicios', index=False)
                    
            # Check Products sheet
            df_products = pd.read_excel(DB_FILE, sheet_name='Productos')
            cols_to_add = []
            if 'Metodo_Pago' not in df_products.columns: cols_to_add.append('Metodo_Pago')
            if 'Marca' not in df_products.columns: cols_to_add.append('Marca')
            if 'Descripcion' not in df_products.columns: cols_to_add.append('Descripcion')
            
            needs_update = False
            if cols_to_add:
                print(f"Migrating Products sheet: Adding {cols_to_add}...")
                for col in cols_to_add:
                    df_products[col] = '' if col != 'Metodo_Pago' else 'Efectivo'
                needs_update = True
                
            if 'Sede' not in df_products.columns:
                print("Migrating Products sheet: Adding Sede column...")
                df_products.insert(0, 'Sede', 'Principal')
                needs_update = True
                
            if needs_update:
                with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df_products.to_excel(writer, sheet_name='Productos', index=False)
            
            # Check Gastos sheet
            try:
                df_gastos = pd.read_excel(DB_FILE, sheet_name='Gastos')
                if 'Sede' not in df_gastos.columns:
                    print("Migrating Gastos sheet: Adding Sede column...")
                    df_gastos.insert(0, 'Sede', 'Principal')
                    with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_gastos.to_excel(writer, sheet_name='Gastos', index=False)
            except:
                pass
            
            # Check Inventory sheet
            try:
                df_inventario = pd.read_excel(DB_FILE, sheet_name='Inventario')
                needs_update = False
                
                if 'Marca' not in df_inventario.columns:
                    df_inventario['Marca'] = ''
                    needs_update = True
                if 'Estado' not in df_inventario.columns:
                    df_inventario['Estado'] = 'Nuevo'
                    needs_update = True
                if 'Sede' not in df_inventario.columns:
                    print("Migrating Inventario sheet: Adding Sede column...")
                    df_inventario.insert(0, 'Sede', 'Principal')
                    needs_update = True
                    
                if needs_update:
                    with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_inventario.to_excel(writer, sheet_name='Inventario', index=False)
            except ValueError:
                print("Migrating: Adding Inventario sheet...")
                with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='a') as writer:
                    pd.DataFrame(columns=['Sede', 'Producto', 'Marca', 'Descripcion', 'Cantidad', 'Unidad', 'Valor', 'Estado', 'Fecha_Actualizacion']).to_excel(writer, sheet_name='Inventario', index=False)

            # Check Citas sheet
            wb = openpyxl.load_workbook(DB_FILE)
            if 'Citas' not in wb.sheetnames:
                ws = wb.create_sheet('Citas')
                ws.append(['ID', 'Sede', 'Fecha', 'Hora', 'Cliente', 'Telefono', 'Servicio', 'Notas', 'Estado'])
                wb.save(DB_FILE)
                print("Migrated database: Added Citas sheet")
            else:
                ws = wb['Citas']
                headers = [cell.value for cell in ws[1]]
                
                needs_save = False
                if 'ID' not in headers:
                    ws.insert_cols(1)
                    ws.cell(row=1, column=1, value='ID')
                    import uuid
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=1, value=str(uuid.uuid4()))
                    headers = [cell.value for cell in ws[1]]  # Refresh headers
                    needs_save = True
                    
                if 'Sede' not in headers:
                    print("Migrating Citas sheet: Adding Sede column...")
                    # Insert Sede after ID (column 2)
                    ws.insert_cols(2)
                    ws.cell(row=1, column=2, value='Sede')
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=2, value='Principal')
                    needs_save = True
                    
                if needs_save:
                    wb.save(DB_FILE)
                    print("Migrated database: Updated Citas sheet")

            # Check GastosMensuales sheet
            try:
                wb = openpyxl.load_workbook(DB_FILE)
                if 'GastosMensuales' not in wb.sheetnames:
                    ws = wb.create_sheet('GastosMensuales')
                    ws.append(['Sede', 'Mes', 'Tipo', 'Valor', 'Fecha_Registro'])
                    wb.save(DB_FILE)
                    print("Migrated database: Added GastosMensuales sheet")
                else:
                    # Check columns if needed, strict migration might not be needed if we just created it
                    pass
            except Exception as e:
                print(f"Error checking GastosMensuales: {e}")

        except Exception as e:
            print(f"Error during migration: {e}")

def init_db():
    """Initialize the Excel database if it doesn't exist."""
    if not os.path.exists(DB_FILE):
        with pd.ExcelWriter(DB_FILE, engine='openpyxl') as writer:
            pd.DataFrame(columns=['Sede', 'Fecha', 'Estilista', 'Servicio', 'Valor', 'Comision', 'Metodo_Pago']).to_excel(writer, sheet_name='Servicios', index=False)
            pd.DataFrame(columns=['Sede', 'Fecha', 'Estilista', 'Producto', 'Marca', 'Descripcion', 'Valor', 'Comision', 'Metodo_Pago']).to_excel(writer, sheet_name='Productos', index=False)
            pd.DataFrame(columns=['Sede', 'Fecha', 'Descripcion', 'Valor']).to_excel(writer, sheet_name='Gastos', index=False)
            pd.DataFrame(columns=['Sede', 'Producto', 'Marca', 'Descripcion', 'Cantidad', 'Unidad', 'Valor', 'Estado', 'Fecha_Actualizacion']).to_excel(writer, sheet_name='Inventario', index=False)
            pd.DataFrame(columns=['ID', 'Sede', 'Fecha', 'Hora', 'Cliente', 'Telefono', 'Servicio', 'Notas', 'Estado']).to_excel(writer, sheet_name='Citas', index=False)
            pd.DataFrame(columns=['Sede', 'Mes', 'Tipo', 'Valor', 'Fecha_Registro']).to_excel(writer, sheet_name='GastosMensuales', index=False)
        print(f"Database {DB_FILE} created.")
    else:
        check_and_migrate_db()


def get_stylists():
    try:
        with open(STYLISTS_FILE, 'r') as f:
            return json.load(f)
    except:
        return []

def save_stylists(stylists):
    with open(STYLISTS_FILE, 'w') as f:
        json.dump(stylists, f)

def get_services():
    try:
        with open(SERVICES_FILE, 'r') as f:
            data = json.load(f)
            # Migration check: if list of strings, convert to objects
            if data and isinstance(data[0], str):
                new_data = [{'name': s, 'value': 0} for s in data]
                save_services(new_data)
                return new_data
            return data
    except:
        return []

def save_services(services):
    with open(SERVICES_FILE, 'w') as f:
        json.dump(services, f)

def get_sedes():
    try:
        with open(SEDES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        # Return default sede if file doesn't exist
        default = ['Principal']
        save_sedes(default)
        return default

def save_sedes(sedes):
    with open(SEDES_FILE, 'w', encoding='utf-8') as f:
        json.dump(sedes, f, ensure_ascii=False, indent=2)

def get_default_sede():
    try:
        sedes = get_sedes()
        if sedes and len(sedes) > 0:
            return sedes[0]
        return 'Bolivia' # Fallback
    except:
        return 'Bolivia'

def append_to_excel(sheet_name, data):
    """Append a row of data to the specified sheet in the Excel file."""
    try:
        # Load existing file
        with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Load the sheet to find the last row and get column order
            try:
                reader = pd.read_excel(DB_FILE, sheet_name=sheet_name)
                start_row = len(reader) + 1
                # Get existing column order
                columns = reader.columns.tolist()
            except ValueError:
                start_row = 0
                # If sheet doesn't exist, use data keys as columns
                columns = list(data.keys())
            
            # Create DataFrame with correct column order
            df_new = pd.DataFrame([data], columns=columns)
            
            # Write the new data
            df_new.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
        return True
    except Exception as e:
        print(f"Error writing to Excel: {e}")
        return False

@app.route('/')
@login_required
def index():
    stylists = get_stylists()
    services = get_services()
    sedes = get_sedes()
    return render_template('index.html', stylists=stylists, services=services, sedes=sedes)

@app.route('/certificado')
@login_required
def view_certificate():
    return render_template('certificado.html')

@app.route('/certificado/descargar')
@login_required
def download_certificate_pdf():
    try:
        context = {
            'date': datetime.now().strftime('%d de %B de %Y'),
            'generation_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # Translate month names if possible or use simple format
        months = {
            'January': 'Enero', 'February': 'Febrero', 'March': 'Marzo', 'April': 'Abril',
            'May': 'Mayo', 'June': 'Junio', 'July': 'Julio', 'August': 'Agosto',
            'September': 'Septiembre', 'October': 'Octubre', 'November': 'Noviembre', 'December': 'Diciembre'
        }
        now = datetime.now()
        date_str = f"{now.day} de {months.get(now.strftime('%B'), now.strftime('%B'))} de {now.year}"
        context['date'] = date_str

        pdf = render_pdf('certificado_pdf.html', context)
        if pdf:
            response = make_response(pdf)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = 'attachment; filename=Certificado_Propiedad_MagicalHair.pdf'
            return response
        
        return "Error al generar el PDF del certificado", 500
    except Exception as e:
        return str(e), 500

@app.route('/api/stylist', methods=['POST'])
@login_required
def add_stylist():
    data = request.json
    new_stylist = data.get('name')
    if new_stylist:
        stylists = get_stylists()
        if new_stylist not in stylists:
            stylists.append(new_stylist)
            save_stylists(stylists)
            return jsonify({'status': 'success', 'message': 'Estilista agregado'})
        return jsonify({'status': 'error', 'message': 'El estilista ya existe'})
    return jsonify({'status': 'error', 'message': 'Nombre inválido'})

@app.route('/api/stylist', methods=['DELETE'])
@login_required
def delete_stylist():
    data = request.json
    name_to_delete = data.get('name')
    stylists = get_stylists()
    if name_to_delete in stylists:
        stylists.remove(name_to_delete)
        save_stylists(stylists)
        return jsonify({'status': 'success', 'message': 'Estilista eliminado'})
    return jsonify({'status': 'error', 'message': 'Estilista no encontrado'})

@app.route('/api/service-item', methods=['POST'])
@login_required
def add_service_item():
    data = request.json
    new_service_name = data.get('name')
    new_service_value = float(data.get('value', 0))
    
    if new_service_name:
        services = get_services()
        # Check if exists by name
        if not any(s['name'] == new_service_name for s in services):
            services.append({'name': new_service_name, 'value': new_service_value})
            save_services(services)
            return jsonify({'status': 'success', 'message': 'Servicio agregado'})
        return jsonify({'status': 'error', 'message': 'El servicio ya existe'})
    return jsonify({'status': 'error', 'message': 'Nombre inválido'})

@app.route('/api/service-item', methods=['DELETE'])
@login_required
def delete_service_item():
    data = request.json
    name_to_delete = data.get('name')
    services = get_services()
    
    # Filter out the service with the given name
    new_services = [s for s in services if s['name'] != name_to_delete]
    
    if len(new_services) < len(services):
        save_services(new_services)
        return jsonify({'status': 'success', 'message': 'Servicio eliminado'})
    return jsonify({'status': 'error', 'message': 'Servicio no encontrado'})

@app.route('/api/sedes', methods=['GET'])
@login_required
def get_sedes_api():
    sedes = get_sedes()
    return jsonify({'status': 'success', 'data': sedes})

@app.route('/api/sede', methods=['POST'])
@login_required
def add_sede():
    data = request.json
    new_sede = data.get('name')
    if new_sede:
        sedes = get_sedes()
        if new_sede not in sedes:
            sedes.append(new_sede)
            save_sedes(sedes)
            return jsonify({'status': 'success', 'message': 'Sede agregada'})
        return jsonify({'status': 'error', 'message': 'La sede ya existe'})
    return jsonify({'status': 'error', 'message': 'Nombre inválido'})

@app.route('/api/sede', methods=['DELETE'])
@login_required
def delete_sede():
    data = request.json
    name_to_delete = data.get('name')
    sedes = get_sedes()
    if len(sedes) <= 1:
        return jsonify({'status': 'error', 'message': 'No se puede eliminar la última sede'})
    if name_to_delete in sedes:
        sedes.remove(name_to_delete)
        save_sedes(sedes)
        return jsonify({'status': 'success', 'message': 'Sede eliminada'})
    return jsonify({'status': 'error', 'message': 'Sede no encontrada'})

def calculate_commission(stylist, service, value):
    service_lower = service.lower()
    is_special_service = 'tinte' in service_lower or 'mechas' in service_lower or 'keratina' in service_lower
    
    if 'monica' in stylist.lower():
        if is_special_service:
            return value * 0.50
        return value * 0.40
    
    if 'elizabeth' in stylist.lower():
        if is_special_service:
            return value * 0.60
        return value * 0.50
        
    # Default for others
    return value * 0.50

@app.route('/api/service', methods=['POST'])
@login_required
def add_service():
    data = request.json
    valor = float(data['valor'])
    stylist = data['estilista']
    service = data['servicio']
    metodo_pago = data.get('metodo_pago', 'Efectivo')
    sede = data.get('sede', get_default_sede())
    
    comision = calculate_commission(stylist, service, valor)
    
    record = {
        'Sede': sede,
        'Fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'Estilista': stylist,
        'Servicio': service,
        'Valor': valor,
        'Comision': comision,
        'Metodo_Pago': metodo_pago
    }
    
    if append_to_excel('Servicios', record):
        return jsonify({'status': 'success', 'message': 'Servicio registrado', 'comision': comision})
    return jsonify({'status': 'error', 'message': 'Error al guardar'}), 500

@app.route('/api/product', methods=['POST'])
@login_required
def add_product():
    data = request.json
    valor = float(data['valor'])
    metodo_pago = data.get('metodo_pago', 'Efectivo')
    marca = data.get('marca', '')
    descripcion = data.get('descripcion', '')
    sede = data.get('sede', get_default_sede())
    
    record = {
        'Sede': sede,
        'Fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'Estilista': data['estilista'],
        'Producto': data['producto'],
        'Marca': marca,
        'Descripcion': descripcion,
        'Valor': valor,
        'Comision': valor * 0.10,
        'Metodo_Pago': metodo_pago
    }
    
    if append_to_excel('Productos', record):
        # Update Inventory
        try:
            df_inv = pd.read_excel(DB_FILE, sheet_name='Inventario')
            product_name = data['producto']
            
            # Filter matches by sede first
            matches = df_inv[(df_inv['Producto'] == product_name) & (df_inv['Sede'] == sede)]
            
            # If marca is provided, try to filter by it too
            if marca and 'Marca' in df_inv.columns:
                marca_matches = matches[matches['Marca'] == marca]
                if not marca_matches.empty:
                    matches = marca_matches
            
            if not matches.empty:
                # Take the first match
                idx = matches.index[0]
                current_qty = float(df_inv.at[idx, 'Cantidad'])
                
                if current_qty > 0:
                    new_qty = current_qty - 1
                    df_inv.at[idx, 'Cantidad'] = new_qty
                    
                    if new_qty <= 0:
                        df_inv.at[idx, 'Estado'] = 'Agotado'
                    
                    with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_inv.to_excel(writer, sheet_name='Inventario', index=False)
                        
        except Exception as e:
            print(f"Error updating inventory: {e}")
            
        return jsonify({'status': 'success', 'message': 'Producto registrado y descontado del inventario'})
    return jsonify({'status': 'error', 'message': 'Error al guardar'}), 500

@app.route('/api/expense', methods=['POST'])
@login_required
def add_expense():
    data = request.json
    sede = data.get('sede', get_default_sede())
    record = {
        'Sede': sede,
        'Fecha': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'Descripcion': data['descripcion'],
        'Valor': float(data['valor'])
    }
    
    if append_to_excel('Gastos', record):
        return jsonify({'status': 'success', 'message': 'Gasto registrado'})
    return jsonify({'status': 'error', 'message': 'Error al guardar'}), 500

@app.route('/api/inventory', methods=['GET'])
@login_required
def get_inventory():
    try:
        sede_filter = request.args.get('sede', get_default_sede())
        df = pd.read_excel(DB_FILE, sheet_name='Inventario')
        # Filter by sede
        df = df[df['Sede'] == sede_filter]
        # Replace NaN with empty string or 0 for JSON serialization
        df = df.fillna('')
        inventory = df.to_dict(orient='records')
        return jsonify({'status': 'success', 'data': inventory})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/inventory', methods=['POST'])
@login_required
def add_inventory_item():
    try:
        data = request.json
        producto = data.get('producto')
        marca = data.get('marca', '')
        descripcion = data.get('descripcion', '')
        cantidad = float(data.get('cantidad', 0))
        unidad = data.get('unidad')
        valor = float(data.get('valor', 0))
        estado = data.get('estado', 'Nuevo')
        sede = data.get('sede', get_default_sede())
        
        if not producto:
            return jsonify({'status': 'error', 'message': 'Nombre del producto requerido'}), 400

        # Load existing inventory
        try:
            df = pd.read_excel(DB_FILE, sheet_name='Inventario')
        except ValueError:
            # Sheet might not exist if migration failed or file is weird, but init_db should handle it.
            df = pd.DataFrame(columns=['Producto', 'Marca', 'Descripcion', 'Cantidad', 'Unidad', 'Valor', 'Estado', 'Fecha_Actualizacion'])

        # Check if product exists to update
        # Ensure columns exist and fill NaNs for comparison
        if 'Marca' not in df.columns: df['Marca'] = ''
        if 'Descripcion' not in df.columns: df['Descripcion'] = ''
        
        df['Marca'] = df['Marca'].fillna('')
        df['Descripcion'] = df['Descripcion'].fillna('')
        
        # Check for exact match on Sede, Name, Brand, and Description
        mask = (df['Sede'] == sede) & (df['Producto'] == producto) & (df['Marca'] == marca) & (df['Descripcion'] == descripcion)
        
        if mask.any():
            # Update existing
            idx = df.index[mask].tolist()[0]
            df.at[idx, 'Cantidad'] = cantidad
            df.at[idx, 'Unidad'] = unidad
            df.at[idx, 'Valor'] = valor
            df.at[idx, 'Estado'] = estado
            df.at[idx, 'Fecha_Actualizacion'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            message = 'Producto actualizado'
        else:
            # Add new
            new_row = {
                'Sede': sede,
                'Producto': producto,
                'Marca': marca,
                'Descripcion': descripcion,
                'Cantidad': cantidad,
                'Unidad': unidad,
                'Valor': valor,
                'Estado': estado,
                'Fecha_Actualizacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            message = 'Producto agregado al inventario'
            
        # Save back to Excel
        with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name='Inventario', index=False)
            
        return jsonify({'status': 'success', 'message': message})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/inventory', methods=['DELETE'])
@login_required
def delete_inventory_item():
    try:
        data = request.json
        producto = data.get('producto')
        marca = data.get('marca', '')
        descripcion = data.get('descripcion', '')
        sede = data.get('sede', get_default_sede())
        
        df = pd.read_excel(DB_FILE, sheet_name='Inventario')
        
        # Ensure columns exist
        if 'Marca' not in df.columns: df['Marca'] = ''
        if 'Descripcion' not in df.columns: df['Descripcion'] = ''
        
        df['Marca'] = df['Marca'].fillna('')
        df['Descripcion'] = df['Descripcion'].fillna('')
        
        # Filter by sede, producto, marca, descripcion
        mask = (df['Sede'] == sede) & (df['Producto'] == producto) & (df['Marca'] == marca) & (df['Descripcion'] == descripcion)
        
        if mask.any():
            df = df[~mask]
            with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Inventario', index=False)
            return jsonify({'status': 'success', 'message': 'Producto eliminado del inventario'})
        return jsonify({'status': 'error', 'message': 'Producto no encontrado'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/summary', methods=['GET'])
@login_required
def get_summary():
    try:
        # Get date and sede from query parameters
        # Get date and sede from query parameters
        date_filter = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        sede_filter = request.args.get('sede', get_default_sede())
        summary_data = []
        total_valor = 0
        total_comision = 0
        total_gastos = 0

        # Read Services
        try:
            df_services = pd.read_excel(DB_FILE, sheet_name='Servicios')
            for idx, row in df_services.iterrows():
                if str(row['Fecha']).startswith(date_filter) and row['Sede'] == sede_filter:
                    val = float(row['Valor']) if pd.notna(row['Valor']) else 0.0
                    com = float(row['Comision']) if pd.notna(row['Comision']) else 0.0
                    
                    metodo = row.get('Metodo_Pago', 'N/A')
                    if pd.isna(metodo):
                        metodo = 'N/A'
                        
                    summary_data.append({
                        'id': int(idx),
                        'sheet': 'Servicios',
                        'estilista': row['Estilista'],
                        'descripcion': row['Servicio'],
                        'valor': val,
                        'comision': com,
                        'tipo': 'Servicio',
                        'metodo_pago': metodo
                    })
                    total_valor += val
                    total_comision += com
        except Exception as e:
            print(f"Error reading services: {e}")

        # Read Products
        try:
            df_products = pd.read_excel(DB_FILE, sheet_name='Productos')
            for idx, row in df_products.iterrows():
                if str(row['Fecha']).startswith(date_filter) and row['Sede'] == sede_filter:
                    val = float(row['Valor']) if pd.notna(row['Valor']) else 0.0
                    com = float(row['Comision']) if pd.notna(row['Comision']) else 0.0
                    
                    metodo = row.get('Metodo_Pago', 'N/A')
                    if pd.isna(metodo):
                        metodo = 'N/A'

                    summary_data.append({
                        'id': int(idx),
                        'sheet': 'Productos',
                        'estilista': row['Estilista'],
                        'descripcion': row['Producto'],
                        'valor': val,
                        'comision': com,
                        'tipo': 'Producto',
                        'metodo_pago': metodo
                    })
                    total_valor += val
                    total_comision += com
        except Exception as e:
            print(f"Error reading products: {e}")

        # Read Expenses
        try:
            df_expenses = pd.read_excel(DB_FILE, sheet_name='Gastos')
            for _, row in df_expenses.iterrows():
                if str(row['Fecha']).startswith(date_filter) and row['Sede'] == sede_filter:
                    val = float(row['Valor']) if pd.notna(row['Valor']) else 0.0
                    total_gastos += val
        except Exception as e:
            print(f"Error reading expenses: {e}")

        # Calculate profit: Net Sales - Expenses - Commissions
        utilidad = total_valor - total_gastos - total_comision

        return jsonify({
            'status': 'success',
            'data': summary_data,
            'totals': {
                'valor': total_valor,
                'comision': total_comision,
                'gastos': total_gastos,
                'utilidad': utilidad
            }
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/summary/update', methods=['POST'])
@login_required
def update_summary_item():
    try:
        data = request.json
        sheet_name = data.get('sheet')
        try:
            row_id = int(data.get('id'))
            new_valor = float(data.get('valor'))
            new_comision = float(data.get('comision'))
        except (ValueError, TypeError):
             return jsonify({'status': 'error', 'message': 'Invalid numeric data'}), 400
        
        if sheet_name not in ['Servicios', 'Productos']:
            return jsonify({'status': 'error', 'message': 'Invalid sheet'}), 400
            
        # Read the dataframe
        df = pd.read_excel(DB_FILE, sheet_name=sheet_name)
        
        # Verify index exists
        if row_id not in df.index:
            return jsonify({'status': 'error', 'message': 'Item not found'}), 404
            
        # Update values
        df.at[row_id, 'Valor'] = new_valor
        df.at[row_id, 'Comision'] = new_comision
        
        # Save back
        with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
        return jsonify({'status': 'success', 'message': 'Item actualizado'})
            
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

def render_pdf(template_src, context_dict):
    template = app.jinja_env.get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return result.getvalue()
    return None

@app.route('/export_pdf')
@login_required
def export_pdf():
    try:
        date_filter = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
        
        # Reuse logic to get data (this could be refactored into a separate function to avoid duplication)
        # For now, I'll just call the internal logic or copy it. 
        # Calling the internal logic is cleaner if I extract it, but to minimize changes I will copy the core logic 
        # or better yet, I will refactor get_summary to return a dict that I can use here.
        
        # Let's do a quick extraction of the data gathering logic to avoid code duplication? 
        # Actually, for safety and speed, I will duplicate the data gathering logic here slightly modified for Python usage 
        # (not returning jsonify) or just call the same logic.
        
        summary_data = []
        total_valor = 0
        total_comision = 0
        total_gastos = 0

        # Read Services
        try:
            df_services = pd.read_excel(DB_FILE, sheet_name='Servicios')
            for _, row in df_services.iterrows():
                if str(row['Fecha']).startswith(date_filter):
                    val = float(row['Valor']) if pd.notna(row['Valor']) else 0.0
                    com = float(row['Comision']) if pd.notna(row['Comision']) else 0.0
                    metodo = row.get('Metodo_Pago', 'N/A')
                    if pd.isna(metodo): metodo = 'N/A'
                        
                    summary_data.append({
                        'estilista': row['Estilista'],
                        'descripcion': row['Servicio'],
                        'valor': val,
                        'comision': com,
                        'tipo': 'Servicio',
                        'metodo_pago': metodo
                    })
                    total_valor += val
                    total_comision += com
        except: pass

        # Read Products
        try:
            df_products = pd.read_excel(DB_FILE, sheet_name='Productos')
            for _, row in df_products.iterrows():
                if str(row['Fecha']).startswith(date_filter):
                    val = float(row['Valor']) if pd.notna(row['Valor']) else 0.0
                    com = float(row['Comision']) if pd.notna(row['Comision']) else 0.0
                    metodo = row.get('Metodo_Pago', 'N/A')
                    if pd.isna(metodo): metodo = 'N/A'

                    summary_data.append({
                        'estilista': row['Estilista'],
                        'descripcion': row['Producto'],
                        'valor': val,
                        'comision': com,
                        'tipo': 'Producto',
                        'metodo_pago': metodo
                    })
                    total_valor += val
                    total_comision += com
        except: pass

        # Read Expenses
        try:
            df_expenses = pd.read_excel(DB_FILE, sheet_name='Gastos')
            for _, row in df_expenses.iterrows():
                if str(row['Fecha']).startswith(date_filter):
                    val = float(row['Valor']) if pd.notna(row['Valor']) else 0.0
                    total_gastos += val
        except: pass

        utilidad = total_valor - total_gastos - total_comision
        
        totals = {
            'valor': total_valor,
            'comision': total_comision,
            'gastos': total_gastos,
            'utilidad': utilidad
        }

        # Logo path
        logo_path = os.path.join(app.root_path, 'Logo', 'Magical_Hair.png')
        # Check if exists, if not try .ico or None
        if not os.path.exists(logo_path):
             logo_path = os.path.join(app.root_path, 'Logo', 'Magical_Hair.ico')
             if not os.path.exists(logo_path):
                 logo_path = None

        context = {
            'data': summary_data,
            'totals': totals,
            'date': date_filter,
            'logo_path': logo_path,
            'generation_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        pdf = render_pdf('pdf_report.html', context)
        if pdf:
            response = make_response(pdf)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'attachment; filename=Cierre_{date_filter}.pdf'
            return response
        
        return "Error generating PDF", 500

    except Exception as e:
        return str(e), 500

@app.route('/api/statistics', methods=['GET'])
@login_required
def get_statistics():
    try:
        # Get month from query parameter (format: YYYY-MM)
        month_filter = request.args.get('month', datetime.now().strftime('%Y-%m'))
        year, month = month_filter.split('-')
        sede_filter = request.args.get('sede', '') # Optional sede filter
        
        # Read all sheets
        servicios_df = pd.read_excel(DB_FILE, sheet_name='Servicios')
        productos_df = pd.read_excel(DB_FILE, sheet_name='Productos')
        gastos_df = pd.read_excel(DB_FILE, sheet_name='Gastos')
        inventario_df = pd.read_excel(DB_FILE, sheet_name='Inventario')
        
        try:
            gastos_mensuales_df = pd.read_excel(DB_FILE, sheet_name='GastosMensuales')
        except ValueError:
            gastos_mensuales_df = pd.DataFrame(columns=['Sede', 'Mes', 'Tipo', 'Valor'])
        
        # Apply Sede Filter if provided
        if sede_filter:
             servicios_df = servicios_df[servicios_df['Sede'] == sede_filter]
             productos_df = productos_df[productos_df['Sede'] == sede_filter]
             gastos_df = gastos_df[gastos_df['Sede'] == sede_filter]
             inventario_df = inventario_df[inventario_df['Sede'] == sede_filter]
             gastos_mensuales_df = gastos_mensuales_df[gastos_mensuales_df['Sede'] == sede_filter]

        # Filter by month
        servicios_df['Fecha'] = pd.to_datetime(servicios_df['Fecha'].astype(str).str.slice(0, 10), errors='coerce')
        productos_df['Fecha'] = pd.to_datetime(productos_df['Fecha'].astype(str).str.slice(0, 10), errors='coerce')
        gastos_df['Fecha'] = pd.to_datetime(gastos_df['Fecha'].astype(str).str.slice(0, 10), errors='coerce')
        
        servicios_month = servicios_df[(servicios_df['Fecha'].dt.year == int(year)) & 
                                       (servicios_df['Fecha'].dt.month == int(month))]
        productos_month = productos_df[(productos_df['Fecha'].dt.year == int(year)) & 
                                       (productos_df['Fecha'].dt.month == int(month))]
        gastos_month = gastos_df[(gastos_df['Fecha'].dt.year == int(year)) & 
                                 (gastos_df['Fecha'].dt.month == int(month))]
        
        # Gastos Mensuales (Fixed)
        gastos_mensuales_month = gastos_mensuales_df[gastos_mensuales_df['Mes'] == month_filter]

        # Calculate totals
        total_ventas = servicios_month['Valor'].sum() + productos_month['Valor'].sum()
        total_gastos = gastos_month['Valor'].sum()
        total_nomina = servicios_month['Comision'].sum() + productos_month['Comision'].sum()
        utilidad_operativa = total_ventas - total_gastos - total_nomina
        
        total_gastos_fijos = gastos_mensuales_month['Valor'].sum()
        utilidad_real = utilidad_operativa - total_gastos_fijos
        
        # Nómina por estilista
        nomina_servicios = servicios_month.groupby('Estilista')['Comision'].sum()
        nomina_productos = productos_month.groupby('Estilista')['Comision'].sum()
        nomina_por_estilista = (nomina_servicios.add(nomina_productos, fill_value=0)).to_dict()
        
        # Ventas por estilista
        ventas_servicios = servicios_month.groupby('Estilista')['Valor'].sum()
        ventas_productos = productos_month.groupby('Estilista')['Valor'].sum()
        ventas_por_estilista = (ventas_servicios.add(ventas_productos, fill_value=0)).to_dict()
        
        # Inventario resumido - Agrupado por producto
        inventario_agrupado = {}
        for _, row in inventario_df.iterrows():
            try:
                producto = row['Producto'] if pd.notna(row['Producto']) else ''
                if not producto:
                    continue
                    
                cantidad = float(row['Cantidad']) if pd.notna(row['Cantidad']) else 0.0
                valor = float(row['Valor']) if pd.notna(row['Valor']) else 0.0
                unidad = row['Unidad'] if pd.notna(row['Unidad']) else ''
                
                # Si el producto ya existe, sumar las cantidades y valores
                if producto in inventario_agrupado:
                    inventario_agrupado[producto]['cantidad'] += cantidad
                    inventario_agrupado[producto]['valor_total'] += cantidad * valor
                else:
                    inventario_agrupado[producto] = {
                        'producto': producto,
                        'cantidad': cantidad,
                        'unidad': unidad,
                        'valor_total': cantidad * valor
                    }
            except:
                continue
        
        # Convertir el diccionario a lista
        inventario_resumido = list(inventario_agrupado.values())
            
        # Yearly Sales Timeline - Last 3 years
        meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
                 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        
        timeline_data = {}
        for y in range(int(year) - 2, int(year) + 1):
            ventas_anuales = []
            for m in range(1, 13):
                try:
                    s_month = servicios_df[(servicios_df['Fecha'].dt.year == y) & 
                                           (servicios_df['Fecha'].dt.month == m)]
                    p_month = productos_df[(productos_df['Fecha'].dt.year == y) & 
                                           (productos_df['Fecha'].dt.month == m)]
                    total = s_month['Valor'].sum() + p_month['Valor'].sum()
                    ventas_anuales.append(float(total))
                except:
                    ventas_anuales.append(0.0)
            timeline_data[str(y)] = ventas_anuales
            
        # Top Servicios (Frecuencia)
        top_servicios = servicios_month['Servicio'].value_counts().head(10).to_dict()
        
        # Estado Inventario
        disponibles = 0
        agotados = 0
        for _, row in inventario_df.iterrows():
            try:
                qty = float(row['Cantidad']) if pd.notna(row['Cantidad']) else 0
                if qty > 0:
                    disponibles += 1
                else:
                    agotados += 1
            except:
                pass
        
        estado_inventario = {
            'Disponibles': disponibles,
            'Agotados': agotados
        }
        
        return jsonify({
            'status': 'success',
            'data': {
                'totales': {
                    'ventas': float(total_ventas),
                    'gastos': float(total_gastos),
                    'nomina': float(total_nomina),
                    'utilidad_operativa': float(utilidad_operativa),
                    'gastos_fijos': float(total_gastos_fijos),
                    'utilidad_real': float(utilidad_real)
                },
                'gastos_mensuales_detalle': gastos_mensuales_month[['Tipo', 'Valor']].to_dict(orient='records'),
                'nomina_por_estilista': {k: float(v) for k, v in nomina_por_estilista.items()},
                'ventas_por_estilista': {k: float(v) for k, v in ventas_por_estilista.items()},
                'top_servicios': top_servicios,
                'estado_inventario': estado_inventario,
                'inventario': inventario_resumido,
                'timeline': timeline_data
            }
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/appointment', methods=['POST'])
def add_appointment():
    try:
        import uuid
        data = request.json
        sede = data.get('sede', get_default_sede())
        record = {
            'ID': str(uuid.uuid4()),
            'Sede': sede,
            'Fecha': data['fecha'],
            'Hora': data['hora'],
            'Cliente': data['cliente'],
            'Telefono': data['telefono'],
            'Servicio': data['servicio'],
            'Notas': data.get('notas', ''),
            'Estado': 'Pendiente'
        }
        
        if append_to_excel('Citas', record):
            return jsonify({'status': 'success', 'message': 'Cita agendada correctamente'})
        return jsonify({'status': 'error', 'message': 'Error al guardar la cita'}), 500
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/appointment/<appointment_id>', methods=['DELETE'])
def delete_appointment(appointment_id):
    try:
        wb = openpyxl.load_workbook(DB_FILE)
        ws = wb['Citas']
        
        # Find row by ID (assuming ID is in column 1)
        row_to_delete = None
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=1).value) == appointment_id:
                row_to_delete = row
                break
        
        if row_to_delete:
            ws.delete_rows(row_to_delete)
            wb.save(DB_FILE)
            return jsonify({'status': 'success', 'message': 'Cita eliminada correctamente'})
        else:
            return jsonify({'status': 'error', 'message': 'Cita no encontrada'}), 404
            
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/appointment/<appointment_id>', methods=['PUT'])
def update_appointment(appointment_id):
    try:
        data = request.json
        wb = openpyxl.load_workbook(DB_FILE)
        ws = wb['Citas']
        
        # Find row by ID
        row_to_update = None
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=1).value) == appointment_id:
                row_to_update = row
                break
        
        if row_to_update:
            # Map fields to columns (assuming order: ID, Fecha, Hora, Cliente, Telefono, Servicio, Notas, Estado)
            # 1-based indexing
            if 'fecha' in data: ws.cell(row=row_to_update, column=2, value=data['fecha'])
            if 'hora' in data: ws.cell(row=row_to_update, column=3, value=data['hora'])
            if 'cliente' in data: ws.cell(row=row_to_update, column=4, value=data['cliente'])
            if 'telefono' in data: ws.cell(row=row_to_update, column=5, value=data['telefono'])
            if 'servicio' in data: ws.cell(row=row_to_update, column=6, value=data['servicio'])
            if 'notas' in data: ws.cell(row=row_to_update, column=7, value=data.get('notas', ''))
            
            wb.save(DB_FILE)
            return jsonify({'status': 'success', 'message': 'Cita actualizada correctamente'})
        else:
            return jsonify({'status': 'error', 'message': 'Cita no encontrada'}), 404
            
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/appointments', methods=['GET'])
def get_appointments():
    try:
        date_filter = request.args.get('date')
        sede_filter = request.args.get('sede', get_default_sede())
        
        try:
            df = pd.read_excel(DB_FILE, sheet_name='Citas')
        except ValueError:
            return jsonify({'status': 'success', 'data': []})
            
        df = df.fillna('')
        
        # Robust Sede filtering
        if 'Sede' in df.columns:
            # Convert to string and strip whitespace for robust comparison
            df['Sede'] = df['Sede'].astype(str).str.strip()
            df = df[df['Sede'] == sede_filter]
        
        if date_filter:
            # Robust Date filtering
            # Convert both column and filter to datetime to compare
            try:
                # Convert column to datetime, coerce errors to NaT
                df['Fecha_DT'] = pd.to_datetime(df['Fecha'], errors='coerce')
                
                # Create filter date object
                filter_date = pd.to_datetime(date_filter)
                
                # Filter rows where valid dates match
                # Check where expected format matches YYYY-MM-DD
                df = df[df['Fecha_DT'].dt.strftime('%Y-%m-%d') == filter_date.strftime('%Y-%m-%d')]
            except Exception as e:
                # Fallback to string comparison if datetime conversion fails completely
                df['Fecha'] = df['Fecha'].astype(str).str.strip()
                df = df[df['Fecha'] == date_filter]
            
        # Sort by Time
        try:
            df = df.sort_values('Hora')
        except:
            pass

        # Cleanup temporary column before conversion
        if 'Fecha_DT' in df.columns:
            df = df.drop(columns=['Fecha_DT'])

        appointments = df.to_dict(orient='records')
        return jsonify({'status': 'success', 'data': appointments})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/alerts', methods=['GET'])
def get_alerts():
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        
        try:
            df = pd.read_excel(DB_FILE, sheet_name='Citas')
        except ValueError:
            return jsonify({'status': 'success', 'alerts': []})
            
        df = df.fillna('')
        df['Fecha'] = df['Fecha'].astype(str)
        
        # Filter for today
        today_appointments = df[df['Fecha'] == today]
        
        alerts = []
        if not today_appointments.empty:
            count = len(today_appointments)
            alerts.append({
                'type': 'info',
                'message': f'Tienes {count} cita(s) programada(s) para hoy.'
            })
            
        return jsonify({'status': 'success', 'alerts': alerts})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/monthly-expenses', methods=['POST'])
@login_required
def add_update_monthly_expenses():
    try:
        data = request.json
        sede = data.get('sede', get_default_sede())
        mes = data.get('mes')  # YYYY-MM
        expenses = data.get('expenses', [])  # List of {tipo: '...', valor: ...}
        
        if not mes or not expenses:
             return jsonify({'status': 'error', 'message': 'Datos incompletos'}), 400

        # Load existing data
        try:
            df = pd.read_excel(DB_FILE, sheet_name='GastosMensuales')
        except ValueError:
            df = pd.DataFrame(columns=['Sede', 'Mes', 'Tipo', 'Valor', 'Fecha_Registro'])

        # Create a timestamp
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # We want to replace existing entries for this Sede + Mes + Tipo, or append if new
        for exp in expenses:
            tipo = exp['tipo']
            valor = float(exp['valor'])
            
            # Check if exists
            mask = (df['Sede'] == sede) & (df['Mes'] == mes) & (df['Tipo'] == tipo)
            
            if mask.any():
                # Update
                idx = df.index[mask].tolist()[0]
                df.at[idx, 'Valor'] = valor
                df.at[idx, 'Fecha_Registro'] = now
            else:
                # Add new
                new_row = {
                    'Sede': sede,
                    'Mes': mes,
                    'Tipo': tipo,
                    'Valor': valor,
                    'Fecha_Registro': now
                }
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

        with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name='GastosMensuales', index=False)
            
        return jsonify({'status': 'success', 'message': 'Gastos mensuales guardados correctamentos'})

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/monthly-expenses', methods=['GET'])
@login_required
def get_monthly_expenses():
    try:
        sede = request.args.get('sede', get_default_sede())
        mes = request.args.get('mes') # YYYY-MM
        
        try:
            df = pd.read_excel(DB_FILE, sheet_name='GastosMensuales')
        except ValueError:
            return jsonify({'status': 'success', 'data': []})
            
        df = df.fillna('')
        
        # Filter
        if mes:
            df = df[(df['Sede'] == sede) & (df['Mes'] == mes)]
        else:
             df = df[df['Sede'] == sede]
             
        data = df.to_dict(orient='records')
        return jsonify({'status': 'success', 'data': data})
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/prediction', methods=['GET'])
@login_required
def get_prediction():
    try:
        sede_filter = request.args.get('sede', get_default_sede())
        
        # Load data
        df_servicios = pd.read_excel(DB_FILE, sheet_name='Servicios')
        df_productos = pd.read_excel(DB_FILE, sheet_name='Productos')
        
        # Filter by sede
        df_servicios = df_servicios[df_servicios['Sede'] == sede_filter]
        df_productos = df_productos[df_productos['Sede'] == sede_filter]
        
        # Convert Fecha to datetime and floor to date
        df_servicios['Fecha'] = pd.to_datetime(df_servicios['Fecha'].astype(str).str.slice(0, 10), errors='coerce').dt.date
        df_productos['Fecha'] = pd.to_datetime(df_productos['Fecha'].astype(str).str.slice(0, 10), errors='coerce').dt.date
        
        # Aggregate by date
        daily_servicios = df_servicios.groupby('Fecha')['Valor'].sum()
        daily_productos = df_productos.groupby('Fecha')['Valor'].sum()
        
        # Combine
        daily_income = daily_servicios.add(daily_productos, fill_value=0).sort_index()
        
        # Convert to DataFrame
        income_df = daily_income.reset_index()
        income_df.columns = ['Fecha', 'Valor']
        
        # Get last 11 months of data
        today = datetime.now().date()
        start_date = (pd.Timestamp(today) - pd.DateOffset(months=11)).date()
        income_df = income_df[income_df['Fecha'] >= start_date]
        
        if income_df.empty:
            return jsonify({'status': 'success', 'historical': [], 'prediction': []})
            
        # Prepare for prediction (Linear Regression)
        # Using ordinal dates as X
        import numpy as np
        from datetime import date
        
        income_df['Ordinal'] = income_df['Fecha'].apply(lambda x: x.toordinal())
        X = income_df['Ordinal'].values.reshape(-1, 1)
        y = income_df['Valor'].values
        
        # Simple Linear Regression: y = mx + b
        if len(income_df) > 1:
            n = len(X)
            sum_x = np.sum(X)
            sum_y = np.sum(y)
            sum_xx = np.sum(X**2)
            sum_xy = np.sum(X * y.reshape(-1, 1))
            
            denominator = (n * sum_xx - sum_x**2)
            if denominator == 0:
                slope = 0
                intercept = np.mean(y)
            else:
                slope = (n * sum_xy - sum_x * sum_y) / denominator
                intercept = (sum_y - slope * sum_x) / n
        else:
            slope = 0
            intercept = y[0] if len(y) > 0 else 0
            
        # Generate next 7 days
        predictions = []
        last_date = income_df['Fecha'].max()
        for i in range(1, 8):
            pred_date = last_date + pd.Timedelta(days=i)
            pred_ordinal = pred_date.toordinal()
            pred_value = max(0, slope * pred_ordinal + intercept)
            predictions.append({
                'fecha': pred_date.strftime('%Y-%m-%d'),
                'valor': float(pred_value)
            })
            
        historical = []
        for _, row in income_df.iterrows():
            historical.append({
                'fecha': row['Fecha'].strftime('%Y-%m-%d'),
                'valor': float(row['Valor'])
            })
            
        return jsonify({
            'status': 'success',
            'historical': historical,
            'prediction': predictions,
            'trend': 'up' if slope > 0 else 'down' if slope < 0 else 'stable'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/revenue-patterns', methods=['GET'])
@login_required
def get_revenue_patterns():
    try:
        sede_filter = request.args.get('sede', get_default_sede())
        
        # Load data
        df_servicios = pd.read_excel(DB_FILE, sheet_name='Servicios')
        df_productos = pd.read_excel(DB_FILE, sheet_name='Productos')
        
        # Filter by sede
        df_servicios = df_servicios[df_servicios['Sede'] == sede_filter]
        df_productos = df_productos[df_productos['Sede'] == sede_filter]
        
        # Convert Fecha
        df_servicios['Fecha'] = pd.to_datetime(df_servicios['Fecha'].astype(str).str.slice(0, 10), errors='coerce').dt.date
        df_productos['Fecha'] = pd.to_datetime(df_productos['Fecha'].astype(str).str.slice(0, 10), errors='coerce').dt.date
        
        # Group by date
        daily_servicios = df_servicios.groupby('Fecha')['Valor'].sum()
        daily_productos = df_productos.groupby('Fecha')['Valor'].sum()
        
        daily_revenue = daily_servicios.add(daily_productos, fill_value=0).sort_index()
        
        # Reset index
        df = daily_revenue.reset_index()
        df.columns = ['Fecha', 'Valor']
        
        # Filter last 11 months
        today = datetime.now().date()
        start_date = (pd.Timestamp(today) - pd.DateOffset(months=11)).date()
        df = df[df['Fecha'] >= start_date]
        
        if df.empty:
            return jsonify({'status': 'success', 'heatmap': [], 'patterns': {}, 'inference': 'Datos insuficientes'})
            
        # 1. Heatmap Data (Week vs Day)
        # We can simulate a heatmap structure: Week Number vs Day of Week
        heatmap_data = []
        for _, row in df.iterrows():
            dt = pd.to_datetime(row['Fecha'])
            heatmap_data.append({
                'date': row['Fecha'].strftime('%Y-%m-%d'),
                'day': dt.day_name(), # English name, we will translate in JS or here
                'day_index': dt.dayofweek, # 0=Mon
                'week': dt.isocalendar()[1],
                'value': float(row['Valor'])
            })
            
        # 2. Average Stats by Day of Week (Inference Basis)
        df['DayOfWeek'] = pd.to_datetime(df['Fecha']).dt.dayofweek
        patterns = df.groupby('DayOfWeek')['Valor'].mean().to_dict()
        
        day_map = {0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves', 4: 'Viernes', 5: 'Sábado', 6: 'Domingo'}
        named_patterns = {day_map[k]: float(v) for k, v in patterns.items()}
        
        # 3. Inference
        sorted_days = sorted(patterns.items(), key=lambda x: x[1], reverse=True)
        if len(sorted_days) >= 2:
            best_day_1 = day_map[sorted_days[0][0]]
            best_day_2 = day_map[sorted_days[1][0]]
            inference = f"Basado en los últimos 11 meses, los días con mayor probabilidad de altos ingresos son los {best_day_1} y {best_day_2}."
        elif len(sorted_days) == 1:
            best_day = day_map[sorted_days[0][0]]
            inference = f"Basado en los últimos 11 meses, el día con mayor probabilidad de altos ingresos es el {best_day}."
        else:
            inference = "No hay suficientes datos para generar una inferencia."
            
        return jsonify({
            'status': 'success', 
            'heatmap': heatmap_data, 
            'patterns': named_patterns, 
            'inference': inference
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/service-demand', methods=['GET'])
@login_required
def get_service_demand():
    try:
        import numpy as np
        sede_filter = request.args.get('sede', get_default_sede())
        
        # Load services data
        try:
            df_servicios = pd.read_excel(DB_FILE, sheet_name='Servicios')
        except ValueError:
             return jsonify({'status': 'success', 'historical': [], 'prediction': [], 'growthService': None})

        # Filter by sede
        if 'Sede' in df_servicios.columns:
            df_servicios = df_servicios[df_servicios['Sede'] == sede_filter]
            if df_servicios.empty:
                # If no data for this sede, use all data
                df_servicios = pd.read_excel(DB_FILE, sheet_name='Servicios')
            
        # Convert Fecha
        df_servicios['Fecha'] = pd.to_datetime(df_servicios['Fecha'].astype(str).str.slice(0, 10), errors='coerce')
        df_servicios = df_servicios.dropna(subset=['Fecha'])
        df_servicios['Fecha'] = df_servicios['Fecha'].dt.date
        
        # Filter last 11 months
        today = datetime.now().date()
        start_date = (pd.Timestamp(today) - pd.DateOffset(months=11)).date()
        df_recent = df_servicios[df_servicios['Fecha'] >= start_date].copy()
        
        if df_recent.empty:
            return jsonify({'status': 'success', 'historical': [], 'prediction': [], 'growthService': None})

        # Categorize Services
        def categorize_service(name):
            name = str(name).lower()
            if 'corte' in name: return 'Corte'
            if 'tinte' in name or 'mechas' in name or 'color' in name or 'iluminaciones' in name or 'keratina' in name: return 'Tintura'
            if 'manicure' in name or 'pedicure' in name or 'uñas' in name or 'semi' in name: return 'Uñas'
            if 'depilacion' in name or 'cejas' in name or 'cera' in name or 'bigote' in name: return 'Depilación'
            return 'Otros'

        df_recent['Tipo'] = df_recent['Servicio'].apply(categorize_service)
        
        # Filter only expected types
        expected = ['Corte', 'Tintura', 'Uñas', 'Depilación']
        df_recent = df_recent[df_recent['Tipo'].isin(expected)]
        
        # Group by Date and Tipo
        daily_counts = df_recent.groupby(['Fecha', 'Tipo']).size().reset_index(name='Count')
        
        # Pivot
        pivot_df = daily_counts.pivot(index='Fecha', columns='Tipo', values='Count').fillna(0).astype(int)
        
        # Ensure all columns exist
        for col in expected:
            if col not in pivot_df.columns:
                pivot_df[col] = 0
                
        # Fill missing dates in range (optional but good for charts)
        # For simplicity, we stick to existing dates or just ensure continuity if needed.
        # Let's reindex to ensure full 60 day range? Maybe overkill, but Chart.js handles gaps if we send labels.
        # But for regression we need continuity or at least correct X values.
        
        # Prepare Historical Data
        historical = []
        # Sort by date
        pivot_df = pivot_df.sort_index()
        
        for date_val, row in pivot_df.iterrows():
            entry = {'fecha': date_val.strftime('%Y-%m-%d')}
            for col in expected:
                entry[col] = int(row[col])
            historical.append(entry)
            
        # Prediction
        predictions = []
        growth_scores = {}
        
        # Create a date range for regression X to be accurate
        # We convert dates to ordinals
        dates = pivot_df.index
        X_dates = np.array([d.toordinal() for d in dates])
        
        last_date = dates.max()
        next_7_days = [last_date + pd.Timedelta(days=i) for i in range(1, 8)]
        next_7_ordinals = np.array([d.toordinal() for d in next_7_days])
        
        for col in expected:
            y = pivot_df[col].values
            X = X_dates.reshape(-1, 1)
            
            if len(y) > 1:
                # Linear Regression
                n = len(X)
                sum_x = np.sum(X)
                sum_y = np.sum(y)
                sum_xx = np.sum(X**2)
                sum_xy = np.sum(X.flatten() * y)
                
                denom = n * sum_xx - sum_x**2
                if denom == 0:
                    slope = 0
                    intercept = np.mean(y)
                else:
                    slope = (n * sum_xy - sum_x * sum_y) / denom
                    intercept = (sum_y - slope * sum_x) / n
            else:
                slope = 0
                intercept = y[0] if len(y) > 0 else 0
                
            # Calculate Total Predicted Volume for next 7 days vs Last 7 days to determine "Growth"
            # Or just use slope? Slope is rate of change per day.
            growth_scores[col] = slope
            
            # Generate predictions
            col_preds = []
            for ordinal in next_7_ordinals:
                val = slope * ordinal + intercept
                col_preds.append(max(0, val)) # No negative services
            
            # Store temporary
            pivot_df[f'{col}_pred'] = col_preds if False else 0 # Just placeholder explanation
            
            # We need to structure 'predictions' as list of dicts: [{fecha:..., Corte:..., ...}, ...]
            # Initialize predictions list if empty
            if not predictions:
                for d in next_7_days:
                    predictions.append({'fecha': d.strftime('%Y-%m-%d')})
            
            for i, val in enumerate(col_preds):
                predictions[i][col] = float(val)

        # Determine highest growth
        # We filter for positive growth only
        positive_growth = {k: v for k, v in growth_scores.items() if v > 0}
        if positive_growth:
            growth_service = max(positive_growth, key=positive_growth.get)
        else:
            # If all are negative or zero, maybe pick the one declining the least? Or None.
            # User wants "resalta cuál servicio crecería más".
            # If nothing grows, maybe none.
            growth_service = None
            if not positive_growth and growth_scores:
                 # Fallback: max slope even if negative (least decline)
                 growth_service = max(growth_scores, key=growth_scores.get)

        return jsonify({
            'status': 'success',
            'historical': historical,
            'prediction': predictions,
            'growthService': growth_service
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', debug=True)
